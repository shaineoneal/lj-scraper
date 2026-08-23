import asyncio
import random
import re
import sys
from pathlib import Path

from rich.panel import Panel
from rich.table import Table

from src.browser import launch_browser_with_fallback

from .config import (
    DEFAULT_USER_DATA_DIR,
    load_config,
    update_status,
)

# Add the parent directory of this file to the Python path if running as a script
if __name__ == "__main__" and not __package__:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    __package__ = "src"

from openpyxl import load_workbook
from playwright.async_api import Error as PlaywrightError
from playwright.async_api import Page, async_playwright

# ==============================================================================
# CSS & STYLING CONSTANTS
# ==============================================================================

CUSTOM_CSS = """
body > article > div.b-singlepost-qrcode.ng-isolate-scope{display: none;}
.b-leaf-footer{display: none;}  
.b-singlepost .i-ljuser B,.b-singlepost-meta-title{font-family:ProximaNova,Tahoma,Arial,sans-serif;font-weight:600}
.b-singlepost{margin:0;padding:30px 0 0;color:#242F33}
.b-singlepost BUTTON:focus,.b-singlepost INPUT:focus,.b-singlepost SELECT:focus,.b-singlepost TEXTAREA:focus{outline-color:#00A3D9}
.b-singlepost-about{margin:0 30px}.b-singlepost-author{margin:0 0 1em;display:-webkit-flex;display:-ms-flexbox;display:flex}
.b-singlepost-author:after{content:"";display:table;border-collapse:collapse;clear:both}
.b-singlepost-author-userpic{float:left;margin:0;padding:0 18px 0 0}
.b-singlepost-author-userpic IMG{border-radius:5px}
.b-singlepost-author-userinfo{margin:0;display:-webkit-flex;display:-ms-flexbox;display:flex;-webkit-flex-direction:column;-ms-flex-direction:column;flex-direction:column}
.b-singlepost-author-userinfo-print{display:none}
.b-singlepost-author-date{font-size:.875rem;letter-spacing:normal;color:#7A9199}
.b-singlepost-author-position{margin:0;color:#7A9199}
.b-singlepost-author-position::before{content:"\00b7\00a0"}
.b-singlepost-author-position a:link,.b-singlepost-author-position a:visited{color:#7A9199}
.b-singlepost-author-position a:active,.b-singlepost-author-position a:hover{color:#00A3D9}
.b-singlepost-author-addfriend{margin-top:auto;padding-top:10px}
.b-singlepost-standout{margin:0;text-align:center;display:none;}
.b-singlepost-wrapper{display:-webkit-flex;display:-ms-flexbox;display:flex;-webkit-flex-direction:column;-ms-flex-direction:column;flex-direction:column;max-width:1100px;margin:0 30px 2.5em}
.b-singlepost-meta{margin:0 0 1.3em;padding:0;list-style:none;font-size:.875rem;letter-spacing:normal}
.b-singlepost-meta-item{margin:0;padding:0;list-style:none}
.b-singlepost-meta-title{text-transform:uppercase;letter-spacing:.05em;color:#839399}
.b-singlepost-meta-title B{font-weight:600}.b-singlepost-meta-data A{margin-right:6px}
.b-singlepost-meta-data A:last-child{margin-left:6px}
.category-panel+.b-singlepost-meta,.category-panel+.copyrighted{margin-top:16px}
.b-singlepost-wrapper .ljlikus-button .svgicon{stroke:#7A9199;stroke-width:2.5px}
.b-singlepost-wrapper .ljlikus--like-active .svgicon{fill:#FF4C44;stroke-width:0}
.b-singlepost-wrapper .ljlikus-action{font-size:14px;font-weight:700}
.b-singlepost-title{display:block;margin:.9em 0 .5em;padding:0;font:300 2.0625em/1 ProximaNova,Helvetica,Arial,sans-serif;color:#242F33}
.b-singlepost-title A:link{color:#242F33}.b-singlepost-title A:visited{color:#5CB7D6}
.b-singlepost-title A:active,.b-singlepost-title A:hover{color:#00A3D9}
.b-singlepost-title .i-posticon-wrapper{display:inline-block;margin:0;padding:0;background-image:url(/air/default_theme/img/sprite.png?v=40848);background-image:url(/air/default_theme/img/sprite.svg?v=40878),none;background-repeat:no-repeat;background-position:0 0}
.b-singlepost-title .i-posticon{visibility:hidden}.b-singlepost-title .i-posticon-delayed{display:-webkit-inline-flex;display:-ms-inline-flexbox;display:inline-flex;vertical-align:-2px}
.b-singlepost-title .i-posticon-delayed .flaticon--hourglass{width:18px;height:26px}
.b-singlepost-title .i-posticon-sticky{width:28px;height:33px;background-position:-160px -320px}
.b-singlepost-title .i-posticon-private{width:22px;height:29px;background-position:0 -320px}
.b-singlepost-title .i-posticon-friendsonly{width:32px;height:34px;background-position:-120px -320px}.b-singlepost-title .i-posticon-protected{width:22px;height:33px;background-position:-40px -320px}HTML BODY .b-singlepost-title .i-ljuser-userhead{vertical-align:baseline!important}
.b-singlepost-body{position:relative;margin:0 0 1.3em}
.b-singlepost-body--delayed .lj-like{display:none!important}
.b-singlepost-body:after{content:"";display:table;border-collapse:collapse;clear:both}.b-singlepost-body FIGURE,.b-singlepost-body P{margin:0 0 1.5em}
.b-singlepost-body BLOCKQUOTE{margin:0 0 1.5em 1.5em;font-style:italic;color:#666}
.b-singlepost-body PRE{white-space:pre;white-space:pre-wrap;word-wrap:break-word;line-height:1.5em}
.b-singlepost-body IMG:not(.emoji){height:auto!important;max-width:100%;max-height:2048px;box-sizing:border-box}
.b-singlepost-body .b-mediaplaceholder{max-width:100%}
.b-singlepost-tags{margin:0 0 .85em;font-size:.875rem;letter-spacing:normal}
.b-singlepost-tags STRONG{margin:0;text-transform:uppercase;letter-spacing:.05em;font-family:ProximaNova,Tahoma,Arial,sans-serif;font-weight:600;color:#839399}
.b-singlepost-withbodybanner .b-singlepost-wrapper{max-width:none}
.b-leaf-article,.b-singlepost-withbodybanner .b-singlepost-body,.b-singlepost-withbodybanner .b-singlepost-bodywrapper .b-singlepost-body,.b-singlepost-withbodybanner .b-singlepost-meta,.b-singlepost-withbodybanner .b-singlepost-tags,.b-singlepost-withbodybanner .b-singlepost-title{max-width:1100px}.b-singlepost-withbodybanner .b-singlepost-bodywrapper{position:relative}
.b-singlepost-postbanner:not(.ljsale--empty){float:right;margin:30px;width:300px}
.b-singlepost-postbanner:first-child{margin-top:0}.b-singlepost-bodybanner{display:none}
.b-singlepost-withbodybanner .b-singlepost-bodybanner{display:none;position:absolute;top:0;right:0;bottom:0;left:1120px}
.b-singlepost-withbodybanner .b-singlepost-bodybanner .allbanners-close{text-align:center}.b-singlepost-withbodybanner .b-singlepost-bodybanner .ljsale__inner{width:300px}
.b-singlepost-withbodybanner .b-singlepost-bodybanner .allbanners DIV[id*=AdFox] DIV:empty,.b-singlepost-withbodybanner .b-singlepost-bodybanner .allbanners DIV[id*=AdFox]:empty{margin-top:-30px}
.b-singlepost-withbodybanner .b-singlepost-bodybanner .allbanners DIV[id*=AdFox] DIV:empty~IFRAME[style*=width][style*=height]{margin-top:30px;margin-left:auto;margin-right:auto}@media all and (min-width:1470px){.b-singlepost-withbodybanner .b-singlepost-bodybanner{display:block}}
.b-singlepost-addfriend-link{display:inline-block;height:28px;line-height:28px;padding:0 12px;border:1px solid silver;border-radius:2px;font-weight:700}.b-singlepost-reactions{display:-webkit-flex;display:-ms-flexbox;display:flex;-webkit-align-items:center;-ms-flex-align:center;align-items:center}.b-singlepost-reactions-likes,[nglj-like-react]{margin-right:20px;display:none;}
.b-singlepost-afterpostbanner{margin:0 30px;display:none;}.b-singlepost-afterpostbanner:not(.ljsale--empty){padding:20px 0;border-top:1px solid #DAE3E6;display:none;}.b-singlepost-afterpostbanner[lj0sale0init*=super_footer]:not(.ljsale--empty){padding:0;border-top:0}
.appwidget-journalpromo{margin:0 0 30px}.b-singlepost-prevnext{position:relative;margin:0 30px;padding:1.5em 0;border-top:1px solid #DAE3E6}.b-singlepost-prevnext-drag{display:none}
.b-singlepost-prevnext-items{margin:0 25px;padding:0;list-style:none}.b-singlepost-prevnext-items:after{content:"";display:table;border-collapse:collapse;clear:both}
.b-singlepost-prevnext-item{position:relative;margin:0;padding:0}.b-singlepost-prevnext-prev{float:left}.b-singlepost-prevnext-next{float:right;text-align:right}
.b-singlepost-prevnext-author{margin:0}.b-singlepost-prevnext-link{display:block;margin:.3em 0 0;font-size:.875rem;letter-spacing:normal}.b-singlepost-prevnext-arrow,.b-singlepost-prevnext-arrow .flaticon{position:absolute;top:0;bottom:0;margin:auto}.b-singlepost-prevnext-arrow{width:25px}
.b-grove,.b-singlepost-prevnext-inner{margin:0}.b-singlepost-prevnext-link:hover .flaticon{fill:currentColor}.b-singlepost-prevnext-arrow-left{left:-25px}
.b-singlepost-prevnext-arrow-right{right:-25px}.b-singlepost-prevnext-arrow-right .flaticon{right:0}

.b-grove .b-tree.svgpreloader:after{position:absolute;top:5em;left:50%;display:none;margin-left:-15px}.b-leaf,.b-tree{position:relative}
.b-grove-switchview{display:none;margin:0;text-align:center}.b-grove-loading .b-grove-switchview,.b-grove-loading .b-tree.svgpreloader:after{display:block}#js .b-grove-loading .b-tree-twig .b-leaf-inner .b-leaf-header .b-leaf-actions,.b-grove-loading .b-tree-twig{visibility:hidden!important}.b-grove-3comments{margin-bottom:24px}
.b-tree{margin:0 30px}.b-leaf{margin:20px 0 10px}.b-leaf::after{content:"";display:table;border-collapse:collapse;clear:both}.b-leaf-inner.svgpreloader:after{position:absolute;top:2px;right:0;display:none}.b-tree-twig:first-child .b-leaf{margin-top:0}.b-leaf-inner{position:relative;min-width:400px}
.b-leaf-header{display:table;position:relative;width:100%;margin:0 0 7px;background:#E3F3FA}.b-leaf-header::after{content:"";display:table;border-collapse:collapse;clear:both}.b-leaf-userpic{display:table-cell;position:relative;width:100px;height:100px;margin:0;border-collapse:collapse;vertical-align:top}
.b-leaf-userpic-inner{display:block;display:table-cell;width:100px;height:100px;margin:0;padding:0;vertical-align:middle;text-align:center}
.b-leaf-userpic IMG{border:0;vertical-align:top}.b-leaf-details{display:table-cell;margin:0;padding:0 0 0 1em;vertical-align:top}
.b-leaf-subject{margin:3px 0 0;font-size:1em;font-weight:700}.b-leaf-subject-link:active,.b-leaf-subject-link:hover,.b-leaf-subject-link:link,.b-leaf-subject-link:visited{pointer-events:none;color:#242F33}.b-leaf-username{margin:7px 0 5px}.i-ljuser-withalias.i-ljuser-showalias .i-ljuser-alias{display:none}
.b-leaf-ipaddr{margin:0 0 0 1em;font-size:.875rem;letter-spacing:normal}.b-leaf-badge{display:none;margin:0 .5em 0 0;padding:3px 10px;border-radius:10px;font:600 .8em/1 ProximaNova,Tahoma,Arial,sans-serif;text-transform:uppercase}
.html-ie11 .b-leaf-badge{padding:4px 10px 0}.b-leaf-badge-best{background:#39BF71;color:#FFF}
.b-leaf-badge-promo{background:#FD8F40;color:#FFF}.b-leaf-best .b-leaf-badge-best,.b-leaf-promo .b-leaf-badge-promo{display:inline}
.b-leaf-meta{display:inline-block;margin:0 1.5em 7px 0;vertical-align:top;font-size:.875rem;letter-spacing:normal}
.b-leaf-createdtime{margin:0}
.b-leaf-shorttime{display:none;margin:0}.b-leaf-editedtime{margin:0 0 0 1em}
.b-leaf-permalink:link,.b-leaf-permalink:visited{color:inherit}
.b-leaf-actions-item A:link,.b-leaf-actions-item A:visited,.b-leaf-permalink:active,.b-leaf-permalink:hover,.b-leaf:hover 
.b-leaf-permalink:link,.b-leaf:hover .b-leaf-permalink:visited{color:#00A3D9}.b-leaf-actions{margin:0;padding:0;list-style:none;text-transform:uppercase;letter-spacing:.05em;font-size:.875rem;font-family:ProximaNova,Tahoma,Arial,sans-serif;font-weight:600}.b-leaf-header .b-leaf-actions{display:inline-block;margin:0 0 7px;vertical-align:top}#js .b-leaf-header .b-leaf-actions{visibility:hidden}#js .b-grove-hover .b-leaf:hover .b-leaf-header .b-leaf-actions,#js .b-leaf-hover .b-leaf-header .b-leaf-actions,.b-tree-promo .b-leaf-actions-cancel_promo{visibility:visible}
.b-leaf-footer .b-leaf-actions{display:-webkit-flex;display:-ms-flexbox;display:flex;-webkit-flex-wrap:wrap;-ms-flex-wrap:wrap;flex-wrap:wrap;float:none}
.b-leaf-footer .b-leaf-actions:after,.b-leaf-footer .b-leaf-actions:before{display:table;border-collapse:collapse;content:""}
.b-leaf-actions:after{clear:both}.b-leaf-actions-item{float:left;margin:0 1.5em 0 0;padding:0;white-space:nowrap}
.b-leaf-actions-item A{display:inline-block;margin:0;padding:0}
.b-leaf-actions-item A:active,.b-leaf-actions-item A:focus,.b-leaf-actions-item A:hover{color:#0086B3}
.b-leaf-actions-item .reaction-stats{margin:-5px 0;display:none;}
.b-leaf-actions-check{color:#00A3D9;display:none;}
.b-leaf-actions-checkbox{margin:0;padding:0;vertical-align:0}.b-leaf-actions-label{cursor:pointer}.b-leaf-actions-label-text{margin:0 0 .1em;pointer-events:none}.b-leaf-actions-label:hover{color:#0086B3}.b-leaf-actions-collapse,.b-leaf-full .b-leaf-footer .b-leaf-actions-check{display:none}#js .b-leaf-actions-collapse{display:block}#js 
.b-leaf-actions-expand,#js .b-leaf-full .b-leaf-footer .b-leaf-actions-collapse,.b-leaf-actions-expand,.b-leaf-actions-new{display:none}
.b-leaf-actions-cancel_promo,.b-leaf-actions-promote{position:absolute;bottom:47px;right:1em;margin:0}.b-tree-promo .b-controls-best{visibility:hidden}.b-leaf-actions-cancel_promo A,.b-leaf-actions-promote A{display:inline;padding:.5em .75em .4em;border:1px solid;border-radius:3px;box-sizing:border-box}.b-leaf-actions-cancel_promo A:link,.b-leaf-actions-cancel_promo A:visited{color:#C00}
.b-leaf-actions-cancel_promo A:active,.b-leaf-actions-cancel_promo A:focus,.b-leaf-actions-cancel_promo A:hover{color:#A60000}
.b-leaf-withsubject .b-leaf-actions-cancel_promo,.b-leaf-withsubject .b-leaf-actions-promote{bottom:34px}
.b-leaf-best.b-leaf-withsubject 
.b-leaf-actions-cancel_promo,.b-leaf-best.b-leaf-withsubject .b-leaf-actions-promote{bottom:38px}.b-leaf-controls{visibility:visible;margin:0;padding:0;display:none}.b-leaf-controls:after,.b-leaf-controls:before{display:table;border-collapse:collapse;content:""}.b-leaf-controls:after{clear:both}#js 
.b-leaf-controls{visibility:hidden}#js .b-grove-hover .b-leaf:hover .b-leaf-controls,#js 
.b-leaf-hover .b-leaf-controls,.b-tree-best .b-leaf .b-controls-cancel_best{visibility:visible}
.b-leaf-controls-item{float:left;width:24px;height:24px;margin:0 7px 0 0;padding:0}
.b-leaf-controls-item-checkbox,.b-leaf-controls-item:empty{display:none}.b-tree-best .b-leaf-actions-promote{visibility:hidden}#js .b-grove-hover .b-leaf-collapsed:hover .b-leaf-actions-check,#js .b-leaf-collapsed .b-leaf-header .b-leaf-actions,#js .b-leaf-hover.b-leaf-collapsed .b-leaf-actions-check{visibility:visible}.b-leaf-article{line-height:1.4}.b-leaf-article::after{content:"";display:table;border-collapse:collapse;clear:both}.b-leaf-article IMG:not(.emoji){height:auto!important;max-width:100%;max-height:2048px}.b-leaf-article EMBED,.b-leaf-article OBJECT{width:100%;max-width:640px}.b-leaf-article IFRAME{max-width:100%}.b-leaf-article IFRAME:-webkit-full-screen{width:auto;max-width:none}
.b-leaf-status{margin:0}.b-leaf-footer{padding:7px 0 0}.b-leaf-withsubject .b-leaf-username{margin:1px 0 2px}.b-leaf-withsubject .b-leaf-header .b-leaf-actions,.b-leaf-withsubject .b-leaf-meta{margin-bottom:3px}.b-leaf-collapsed,.b-tree-twig:first-child .b-leaf .b-leaf-collapsed{display:-webkit-inline-flex;display:-ms-inline-flexbox;display:inline-flex;margin:0 0 6px;padding:0;line-height:1.4em}
.b-leaf-collapsed .b-leaf-inner{position:relative;float:left;display:-webkit-flex;display:-ms-flexbox;display:flex;-webkit-align-items:center;-ms-flex-align:center;align-items:center;padding:2px 0;min-width:150px;white-space:nowrap;-webkit-flex-wrap:wrap;-ms-flex-wrap:wrap;flex-wrap:wrap}.b-leaf-collapsed .b-leaf-createdtime,.b-leaf-collapsed .b-leaf-editedtime,.b-leaf-collapsed .b-leaf-ipaddr,.b-leaf-collapsed .b-leaf-userpic{display:none}
.b-leaf-collapsed .b-leaf-header{width:auto;float:left;margin:0;background:0 0;-webkit-filter:none;filter:none}.b-leaf-collapsed .b-leaf-header:after,.b-leaf-collapsed .b-leaf-header:before{display:none;content:""}.b-leaf-collapsed .b-leaf-details{float:left;padding:0}
.b-leaf-collapsed .b-leaf-subject{float:left;margin:0 .9em 0 0;line-height:18px;font-size:.875rem;font-weight:400}.b-leaf-collapsed .b-leaf-subject-link:active,.b-leaf-collapsed .b-leaf-subject-link:hover,.b-leaf-collapsed .b-leaf-subject-link:link,.b-leaf-collapsed .b-leaf-subject-link:visited{pointer-events:auto}.b-leaf-collapsed .b-leaf-subject-link:link{color:#00A3D9}
.b-leaf-collapsed .b-leaf-subject-link:visited{color:#007399}.b-leaf-collapsed .b-leaf-subject-link:active,.b-leaf-collapsed .b-leaf-subject-link:hover{color:#0086B3}.b-leaf-collapsed .b-leaf-username{float:left;margin:0 .9em 0 0;font-size:1em;line-height:1.2}.b-leaf-collapsed .b-leaf-meta{float:left;margin:0 1em 0 0;line-height:1.3em;color:#000}.b-leaf-collapsed .b-leaf-shorttime{display:inline}
.b-leaf-collapsed .b-leaf-article,.b-leaf-collapsed .b-leaf-controls{display:none}.b-leaf-collapsed .b-leaf-footer{float:left;margin:0;padding:0;line-height:1.3em}.b-leaf-collapsed .b-leaf-actions{display:-webkit-flex;display:-ms-flexbox;display:flex;float:left;margin:0;line-height:1.4em}#js .b-leaf-collapsed .b-leaf-actions-collapse,.b-leaf-collapsed .b-leaf-actions-item{display:none}
.b-leaf-collapsed .b-leaf-actions:after,.b-leaf-collapsed .b-leaf-actions:before{display:none;content:""}
.b-leaf-collapsed .b-leaf-footer-actions .b-leaf-actions-new{-webkit-align-items:center;-ms-flex-align:center;align-items:center}#js .b-leaf-collapsed .b-leaf-actions-expand,#js .b-leaf-collapsed .b-leaf-actions-item-reaction-stats{display:-webkit-flex;display:-ms-flexbox;display:flex;float:left;margin-right:0}
.b-leaf-collapsed .b-leaf-actions-expand A{display:inline}#js .b-leaf-collapsed .b-leaf-actions-item-reaction-stats{margin-left:.5em;margin-right:1em}#js .b-leaf-collapsed .b-leaf-actions-item-reaction-stats[data-reactions-count="0"],#js .b-leaf-collapsed .b-leaf-actions-permalink{display:none}.b-leaf-collapsed .b-leaf-actions-permalink{display:inline;margin-right:0}.b-leaf-collapsed .b-leaf-actions-permalink A{display:inline}.b-leaf-collapsed .b-leaf-actions-check{display:inline;visibility:visible;margin-left:5px;margin-right:1em}
.b-grove-3comments .b-leaf-collapsed .b-leaf-actions-check,.b-leaf-collapsed .b-leaf-actions-label-text{display:none}
.b-leaf-frozen .b-leaf-header,.b-leaf-screened .b-leaf-header,.b-leaf-spammed .b-leaf-header{background:#BDCDD4}
.b-leaf-screened .b-leaf-article,.b-leaf-spammed .b-leaf-article{color:#BFBFBF}
.b-leaf-collapsed.b-leaf-frozen .b-leaf-header,.b-leaf-collapsed.b-leaf-screened .b-leaf-header,.b-leaf-collapsed.b-leaf-spammed .b-leaf-header{background:0 0}.b-leaf-collapsed.b-leaf-screened .b-leaf-meta,.b-leaf-collapsed.b-leaf-screened .b-leaf-username,.b-leaf-collapsed.b-leaf-spammed .b-leaf-meta,.b-leaf-collapsed.b-leaf-spammed .b-leaf-username{color:#CCC}
.b-leaf-clipped{padding:0;margin:0 0 5px}.b-leaf-clipped.b-leaf-collapsed .b-leaf-inner{display:block;white-space:normal}
.b-leaf-clipped .b-leaf-cheader{display:block;padding:7px 12px;background:#EBF0F2}.b-leaf-clipped.b-leaf-selected .b-leaf-cheader{background:#F0F2EB}.b-leaf-clipped.b-leaf-frozen .b-leaf-cheader{background:#BDCDD4}
.b-leaf-clipped .b-leaf-cheader:after,.b-leaf-clipped .b-leaf-cheader:before{display:table;border-collapse:collapse;content:""}
.b-leaf-clipped .b-leaf-cheader:after{clear:both}.b-leaf-clipped .b-leaf-status{float:left;margin:0 20px 0 0;font-weight:400;font-style:italic}
.b-leaf-clipped .b-leaf-controls{float:left;margin:-3px 13px 0 0;clear:none;visibility:hidden}
.b-leaf-clipped.b-leaf-selected .b-leaf-controls,.b-leaf-clipped:hover .b-leaf-controls{visibility:visible}
.b-leaf-clipped .b-leaf-cheader .b-leaf-actions{float:left;visibility:hidden}#js .b-leaf-collapsed.b-leaf-selected .b-leaf-actions-check,#js .b-leaf-deleting .b-leaf-controls,#js .b-leaf-deleting .b-leaf-header .b-leaf-actions,#js .b-leaf-editing .b-leaf-controls,#js .b-leaf-editing .b-leaf-header .b-leaf-actions,#js .b-leaf-expanding.b-leaf-collapsed .b-leaf-actions-check,#js .b-leaf-herbarium .b-leaf-controls,#js .b-leaf-herbarium .b-leaf-header .b-leaf-actions,#js .b-leaf-processing .b-leaf-header .b-leaf-actions,#js .b-leaf-selected .b-leaf-controls,#js .b-leaf-selected .b-leaf-header .b-leaf-actions,.b-leaf-clipped.b-leaf-selected .b-leaf-cheader .b-leaf-actions,.b-leaf-clipped:hover .b-leaf-cheader .b-leaf-actions{visibility:visible}.b-leaf-clipped .b-leaf-actions-check{margin-top:3px}
.b-leaf-clipped.b-leaf-collapsed .b-leaf-footer{margin:0}.b-leaf-editing{margin:0 0 5px}
.b-leaf-editing .b-controls-edit .b-controls-bg{background-position:-10px -758px}.b-leaf-editing .b-leaf-article,.b-leaf-editing .b-leaf-footer{display:none}#js .b-leaf-processing .b-leaf-controls{visibility:visible;padding:0 27px 0 0;background:url(/img/preloader/preloader-blue-blue.gif?v=16423) 100% 50% no-repeat}#js .b-grove-showspam .b-leaf-spammed.b-leaf-processing .b-leaf-controls,#js .b-leaf-frozen.b-leaf-processing .b-leaf-controls,#js .b-leaf-screened.b-leaf-processing .b-leaf-controls{background:url(/img/preloader/preloader-blue-gray.gif?v=16423) 100% 50% no-repeat}#js .b-leaf-selected.b-leaf-processing .b-leaf-controls{background:url(/img/preloader/preloader-blue-yellow.gif?v=16423) 100% 50% no-repeat}#js .b-leaf-clipped.b-leaf-processing .b-leaf-controls{background:url(/img/preloader/preloader-blue-lblue.gif?v=16423) 100% 50% no-repeat}#js .b-leaf-poster.b-leaf-processing .b-leaf-controls{background:url(/img/preloader/preloader-blue-violet.gif?v=36328) 100% 50% no-repeat}.b-leaf-expanding .b-leaf-actions-expandchilds{padding-right:30px;background:url(/img/preloader/preloader-disc-blue-white-16.gif?v=39502) 100% 50% no-repeat}.b-leaf-expanding.b-leaf-collapsed .b-leaf-inner{padding-right:35px}
.b-leaf-expanding.b-leaf-collapsed.b-leaf-selected .b-leaf-inner{background:url(/img/preloader/preloader-blue-yellow.gif?v=16423) 100% 50% no-repeat}.b-leaf-deleting .b-controls-delete .b-controls-bg{background-position:-10px -406px}.b-leaf-collapsed.b-leaf-selected,.b-leaf-collapsed.b-leaf-selected .b-leaf-inner,.b-leaf-selected .b-leaf-header{background:#F0F2EB}.b-leaf-collapsed.b-leaf-selected .b-leaf-header{background:0 0}.b-leaf-modereply{margin:0}.b-leaf-commenting{margin-bottom:20px}.b-leaf-seemore{margin:0 0 20px;padding:0}.b-leaf-seemore .b-leaf-inner{display:inline-block;min-width:0}.b-leaf-seemore-expand,.b-leaf-seemore-from,.b-leaf-seemore-more{margin:0 1em 0 0}.b-leaf-seemore-more{position:relative;font-size:.875rem;letter-spacing:normal}
.b-leaf-seemore-more:after{display:none;content:" ";position:absolute;bottom:-12px;left:20px;z-index:1;font-size:0;line-height:0;width:0;border-top:10px solid transparent;border-top:10px solid rgba(255,255,255,.01);border-bottom:10px solid transparent;border-bottom:10px solid rgba(255,255,255,.01);border-left:10px solid #EBF0F2;-webkit-transform:rotate(22deg);-ms-transform:rotate(22deg);transform:rotate(22deg)}
.b-watering-outer:after,.b-watering-user:after,.b-watering:after,.b-xylem-cells:after,.b-xylem-cells:before,.b-xylem:after,.b-xylem:before{border-collapse:collapse;content:""}.b-watering-fields,.b-watering-user-login-user .b-input{width:100%}
.b-singlepost .b-leaf-seemore-more A,.b-singlepost .b-leaf-seemore-more A:active,.b-singlepost .b-leaf-seemore-more A:hover,.b-singlepost .b-leaf-seemore-more A:link,.b-singlepost .b-leaf-seemore-more A:visited{position:relative;z-index:2;padding:4px 8px;border-bottom:0;border-radius:5px;background:#EBF0F2;color:#00A3D9;text-decoration:none}
.b-leaf-seemore-from{margin-right:.7em}.b-leaf-seemore-users{margin:0 1em 0 0}.b-leaf-seemore-expand{text-transform:uppercase;letter-spacing:.05em;font-family:ProximaNova,sans-serif;font-weight:600;font-size:.875rem}.b-leaf-seemore.b-leaf-expanding .b-leaf-inner{padding-right:20px}.b-leaf-seemore.b-leaf-expanding .b-leaf-inner.svgpreloader:after{display:block}
.b-leaf-cursor{margin-left:-10px;padding-left:7px;border-left:3px solid #FBE5B3}.b-leaf-poster:not(.b-leaf-selected):not(.b-leaf-screened):not(.b-leaf-spammed):not(.b-leaf-frozen) .b-leaf-header,.b-leaf-poster:not(.b-leaf-selected):not(.b-leaf-screened):not(.b-leaf-spammed):not(.b-leaf-frozen).b-leaf-collapsed{background:#E3F3FA}.b-leaf-poster:not(.b-leaf-selected):not(.b-leaf-screened):not(.b-leaf-spammed):not(.b-leaf-frozen)
.b-leaf-collapsed .b-leaf-header{background:0 0}.b-leaf-poster.b-leaf-collapsed .b-leaf-header{padding:0 5px}
.b-leaf-new .b-leaf-footer .b-leaf-actions-new{display:-webkit-inline-flex;display:-ms-inline-flexbox;display:inline-flex;-webkit-align-items:center;-ms-flex-align:center;align-items:center}.b-leaf-collapsed.b-leaf-new .b-leaf-actions-new{display:inline;float:left;margin-right:0;margin-left:1em}
.b-leaf-collapsed.b-leaf-new .b-leaf-footer .b-leaf-actions-new,.b-watering{display:none}
.b-leaf-collapsed.b-leaf-new .b-thisisnew{vertical-align:2px}.b-leaf-new .b-leaf-seemore-more A,.b-leaf-new .b-leaf-seemore-more A:active,.b-leaf-new .b-leaf-seemore-more A:hover,.b-leaf-new .b-leaf-seemore-more A:link,.b-leaf-new .b-leaf-seemore-more A:visited{background:#FFF0A6}.b-watering{overflow:visible!important;position:relative;max-width:970px;margin:0;padding:0;border:0}
.b-watering:after{display:table;clear:both}
.b-watering-show{display:block}
.b-watering-wrapper{overflow:hidden;position:relative;margin:5px 0 30px -20px;padding:20px 35px 20px 20px;border-radius:3px;border:1px solid #DAE3E6;box-sizing:border-box;background:#FFF}.b-msgsystem-errorbox{margin-top:35px}.b-xylem .b-msgsystem-errorbox{margin-top:0}

.b-singlepost-wrapper,.b-tree,.b-xylem{margin-left:16px;margin-right:16px}
.b-xylem{position:relative;margin:0 30px;padding:20px 0 30px;text-align:center}
.b-xylem:after,.b-xylem:before{display:table}.b-xylem:after{clear:both}.b-grove-talkpage .b-xylem{padding-top:0}
.b-xylem-first{border-top:1px solid #DAE3E6}.b-grove-3comments .b-xylem{display:none}.b-grove-3comments .b-xylem-first{display:block}.b-xylem .b-pseudo,.b-xylem .b-pseudo:hover{text-decoration:none}
.b-xylem-cells{margin:0;padding:0;list-style:none}
.b-xylem-cells:after,.b-xylem-cells:before{display:table}.b-xylem-cells:after{clear:both}
.b-xylem-cell{margin:0}.b-xylem-cell-add{float:left;width:30%;text-align:left}
.b-xylem-cell-amount{float:left;width:40%}.b-grove-showspam .b-xylem-cell-amount,.b-xylem-nocomment 
.b-xylem-cell-amount,.b-xylem.active .b-xylem-cell-amount{float:left;width:40%;padding:0 0 0 30%}.b-xylem-cell-spam{float:right;width:29.9%;text-align:right}.b-xylem-cells+.b-pager{margin-top:1em}
.b-xylem-first.active .b-xylem-cell-add{position:relative;float:none;width:100%}.b-xylem-warning{position:relative;margin:0 0 30px;padding:25px 0;text-align:center}.s-horizon .b-xylem-warning{margin-right:-25px;margin-left:-25px;padding-right:25px;padding-left:25px}.s-lanzelot .b-xylem-warning{margin-right:-26px;margin-left:-26px;padding-right:26px;padding-left:26px}
.b-addcomment:active,.b-addcomment:hover,.b-addcomment:link,.b-addcomment:visited{position:relative;display:inline-block;margin:0;padding:0;border:0;text-decoration:none;text-transform:uppercase;font-family:ProximaNova,sans-serif;font-weight:600;font-size:.875rem;letter-spacing:.05em;color:#00A3D9}
.b-addcomment:active,.b-addcomment:hover{color:#5CB7D6}.b-addcomment-active:active,.b-addcomment-active:hover,.b-addcomment-active:link,.b-addcomment-active:visited,.b-xylem.active .b-addcomment:active,.b-xylem.active .b-addcomment:hover,.b-xylem.active .b-addcomment:link,.b-xylem.active .b-addcomment:visited{position:absolute;top:24px;left:0;z-index:2}
.b-addcomment-inner{margin:0}.b-addcomment-icon{display:none}
.b-spamcomments:active,.b-spamcomments:hover,.b-spamcomments:link,.b-spamcomments:visited{margin:0;padding:0;border:0;color:#00A3D9}
.b-spamcomments:active,.b-spamcomments:hover{color:#0086B3}
.b-spamcomments-active:active,.b-spamcomments-active:hover,.b-spamcomments-active:link,.b-spamcomments-active:visited{display:inline-block;margin:0;padding:0;border:0}
.yadirectwide{margin-bottom:24px;margin-left:21px;margin-right:21px}
.b-singlepost-footerbanner{display:none;margin:0 30px}@media all and (min-width:1500px){.b-singlepost-footerbanner{display:block}}
.b-massaction{padding-bottom:5px}.b-massaction-mobile,
.b-massaction-top{display:none}.b-massaction-top{margin-bottom:30px;border-bottom:1px solid #DAE3E6}
.b-massaction .b-flatbutton{margin:0 0 10px}.b-massaction .b-ljbutton{display:inline;position:static;min-height:0;margin:0;padding:0;vertical-align:baseline}
.b-flatbutton{background:#09C;color:#FFF}
.b-flatbutton[disabled]{opacity:.5;pointer-events:none}
.b-massaction-checkall {display:none;}
.b-pager-pages,.b-pager-pages .b-pager-page{text-align:center;display:inline-block;padding:12px 20px;font-size:26px;}
.b-pager-pages,.b-pager-pages .b-pager-page .A{border: 1px solid #666666; text-decoration: none;}
.b-pager-shortcut{display:none}
.b-pager-next .b-pager-shortcut,.b-pager-nopages .b-pager-pages,.b-pager-nopages .b-pager-prev .b-pager-shortcut,.b-pager-nopages .b-pager-prev BR{display:none}
.b-pager-shortcut,.b-pager-nopages .b-pager-pages,.b-pager-nopages .b-pager-prev .b-pager-shortcut,.b-pager-nopages .b-pager-prev BR{display:none}
.b-pager-shortcut,.b-pager-nopages .b-pager-pages,.b-pager-nopages .b-pager-prev .b-pager-shortcut,.b-pager-nopages .b-pager-prev BR{display:none}
"""

# ==============================================================================
# DECOUPLED BUSINESS LOGIC FUNCTIONS
# ==============================================================================

def extract_urls_from_excel(excel_file_path: Path | str) -> list[str]:
    """
    Loads an Excel workbook and reads target post URLs from Column H (skipping header).

    Args:
        excel_file_path: Path to the target Excel (.xlsx) file.

    Returns:
        list[str]: List of target post URLs.
    """

    wb = load_workbook(excel_file_path)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found in the Excel file.")

    posts: list[str] = [cell.value for cell in ws['H'][1:] if cell.value]

    return posts


class LJPost:
    """
    Represents a LiveJournal post. Handles page navigation, extraction of sections
    (title, author/about, content, footer), and archiving the post as HTML.
    """

    def __init__(self, page: Page, url: str):
        self.page: Page = page
        self.url: str = url
        self.title: str = "No Subject"
        self.html_content = None
        self.page_count = 1


    async def _expand_comments(self) -> None:
        """Expands all comments by clicking 'See More' buttons."""
        more_comments_el = self.page.locator('.b-leaf-seemore-more, .b-leaf-actions-expandchilds')
        while await more_comments_el.count() > 3:
            try:
                await more_comments_el.first.click()
            except Exception:
                break  # Exit if unable to click

    async def load(self, index=1) -> None:
        """Navigates to the post URL."""

        if index == 1:
            await self.page.goto(f"{self.url}?s2id=46580551", wait_until="commit")
            await self.page.wait_for_timeout(5000)

            self.page_count = await self.page.locator("li.b-pager-page").count()/2.0 or 1
        else:
            await self.page.goto(f"{self.url}?s2id=46580551&page={index}", wait_until="networkidle")
            await self.page.wait_for_timeout(5000)

        body_el = self.page.locator("body")
        body = await body_el.inner_html()
        if "This page is not available" in body:
            raise ValueError(f"Post not available: {self.url}")

        await self._expand_comments()
        await self.page.wait_for_timeout(5000)

    async def _extract_title(self) -> str:
        """Extracts the post title, defaulting to 'No Subject' if not found."""
        try:
            await self.page.wait_for_selector(".b-singlepost-about", timeout=5000)
            self.title = await self.page.locator('.b-singlepost-title').inner_text(timeout=5000)
        except PlaywrightError:
            self.title = "No Subject"
        return self.title

    async def _extract_about(self) -> str:
        try:
            loc = self.page.locator('.b-singlepost-about')
            if await loc.count() > 0:
                return await loc.inner_html()
        except Exception:
            pass
        return ""

    async def _extract_post_content(self) -> str:
        try:
            loc = self.page.locator('.b-singlepost-wrapper')
            if await loc.count() > 0:
                return await loc.inner_html()
        except Exception:
            pass
        return ""

    async def _extract_footer(self) -> str:
        try:
            loc = self.page.locator('.b-singlepost-prevnext')
            if await loc.count() > 0:
                footer_el = await loc.inner_html()
                footer_el = re.sub(
                    r'<a href="\S+?/(\d+\.html)".+?"></a>(.+?)<h4 class="b-nav-posts__title">(\w+?) post</h4>\s*<p(.*?)>(.+?)</p>',
                    r'\2<h4 class \"b-nav-posts__title\"><a href="./\1" class="b-nav-posts__link" target="_self">\3 post</a></h4>'
                    r'<p\4><a href="\1" class="b-nav-posts__link" target="_self">\5</a></p>',
                    footer_el)


                if footer_el:
                    return "<div prev-next-nav>" + footer_el + "</div></body>\n</html>"
        except Exception:
            pass
        return ""

    async def _extract_comment_count(self) -> str:
        try:
            loc = self.page.locator('.js-amount').first
            if await loc.count() > 0:
                comment_count = await loc.inner_html()
                return comment_count.strip().split(" ")[0]  # Extract numeric part
        except Exception:
            pass
        return "0"

    async def _extract_comments(self) -> str:
        try:
            loc = self.page.locator('.b-tree-root')
            if await loc.count() > 0:
                return await loc.inner_html()

        except Exception:
            pass
        return ""

    async def _create_header(self) -> str:
        """
        Constructs the HTML header with the extracted title and custom CSS.
        """
        header = f"""<!DOCTYPE html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.title}</title>
    <style>
        {CUSTOM_CSS}
    </style>
</head>
<body>"""
        return header

    async def _extract_pagination(self) -> str:
        try:
            loc = self.page.locator('.b-pager--showpages')
            if await loc.count() > 0:
                return await loc.first.inner_html()
        except Exception:
            pass
        return ""


    async def render_html(self) -> str:
        """
        Extracts all post elements (about, wrapper, footer) and renders them
        using the CUSTOM_CSS stylesheet.
        """
        # Ensure title is extracted
        await self._extract_title()

        comment_count = await self._extract_comment_count()
        post = f"{await self._extract_post_content()} \n<br> <h3>{comment_count} Comment(s)</h3>"

        self.html_content = (await self._create_header() + await self._extract_about() + post
                             + await self._extract_comments()
                             + (await self._extract_pagination() if int(comment_count) > 25 else "")
                             + await self._extract_footer())
        return self.html_content

    async def save_to_file(self, output_dir: Path, filename_index= None) -> Path:
        """
        Renders the post HTML and saves it to a file.
        """
        await self.render_html()

        # Resolve filename
        filename = f'{output_dir.name}-{self.url.rstrip(".html/").split("/")[-1]}'
        if filename_index is not None:
            filename += f"&page={filename_index}.html"
        else:
            filename += ".html"

        print(f"    Saving post to: {output_dir / filename}")
        save_path = Path(output_dir) / f"{filename}"
        with open(save_path, "w", encoding="utf-8") as f:
            f.write(self.html_content)

        return save_path


async def create_post_html(page: Page) -> str | None:
    """
    Constructs a formatted HTML template of a LiveJournal post using target Page selectors.
    (Compatibility wrapper using the LJPost class).

    Args:
        page: Playwright Page instance.

    Returns:
        str | None: Formatted HTML block of the post or None if selector lookup failed.
    """
    try:
        url = page.url
        post = LJPost(page, url)
        return await post.render_html()
    except Exception as e:
        print(f"Error creating post HTML: {e}")
        return None


async def save_posts(page: Page, posts: list[str], output_dir: Path | str, delay: float = 0.0) -> dict:
    """
    Scrapes and archives a list of LiveJournal posts to disk as HTML documents.
    Excel loading is decoupled and handled separately in the caller.

    Args:
        page: Playwright Page instance.
        posts: List of post URLs to scrape.
        output_dir: Target output directory path.
        delay: Seconds to sleep between page fetches to avoid rate limits.

    Returns:
        dict: A summary dictionary (success_count, failed_urls, total).
    """

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    results = {
        "success_count": 0,
        "failed_urls": [],
        "total": len(posts)
    }

    update_status("Saving {len(posts)} post(s)")
    for idx, post_url in enumerate(posts, start=1):
        update_status("Saving post(s)...")
        print(f"[{idx}/{len(posts)}] Scraping post: {post_url}")

        try:
            # Instantiate the LJPost helper
            post = LJPost(page, post_url)

            # Load the post page
            await post.load()
            if delay > 0:
                jitter = delay * random.uniform(0.5, 1.5)
                await asyncio.sleep(jitter)
            if post.page_count <= 1.0:
                await post.save_to_file(output_path)
            else:
                for page_num in range(1, int(post.page_count) + 1):
                    if page_num == 1:
                        await post.save_to_file(output_path, filename_index=page_num)
                    else:
                        await post.load(index=page_num)
                        await post.save_to_file(output_path, filename_index=page_num)

            results["success_count"] += 1

        except Exception as err:
            print(f"Error processing post {post_url}: {err}")
            results["failed_urls"].append(post_url)

    update_status("")
    return results


# ==============================================================================
# CLI RUNNER ENTRY POINT
# ==============================================================================

def parse_url_target(url: str) -> tuple[str, str]:
    """
    Parses a LiveJournal post URL to extract (username, post_filename).
    """
    # Check for pattern https://username.livejournal.com/123.html or similar
    match = re.match(r"https?://([^.]+)\.livejournal\.com/(.*)", url, re.IGNORECASE)
    if match:
        subdomain = re.sub('-', '_', match.group(1))
        path = match.group(2)
        if subdomain not in ("www", "m", "mobile", "classic"):
            filename = path.rstrip("/").split("/")[-1]
            if not filename.endswith(".html"):
                filename = f"{filename}.html"
            return subdomain, filename

    # Check for pattern https://www.livejournal.com/users/username/123.html
    match_users = re.match(r"https?://(?:www\.)?livejournal\.com/users/([^/]+)/(.*)", url, re.IGNORECASE)
    if match_users:
        username = re.sub('-', '_', match_users.group(1))
        path = match_users.group(2)
        filename = path.rstrip("/").split("/")[-1]
        if not filename.endswith(".html"):
            filename = f"{filename}.html"
        return username, filename

    # Fallback
    filename = url.rstrip("/").split("/")[-1]
    if not filename.endswith(".html"):
        filename = f"{filename}.html"
    return "single_posts", filename


async def main_async(target, settings):

    update_status("Initializing...")

    posts = []

    output_dir = Path("save_posts_output")

    if target.endswith(".xlsx"):
        try:
            posts = extract_urls_from_excel(target)
            username = posts[0].lstrip('https://').split('.')[0] if posts else "saved_posts"
            dir_name = re.sub('-', '_', username)
            output_dir = Path("save_posts_output") / dir_name
            print(f"[bold $success]Loaded {len(posts)} posts from Excel file: {target} -> Saving under folder: {output_dir}[/bold $success]")
        except Exception as e:
            print(f"[bold red]Error loading Excel file '{target}': {e}[/bold red]")
            sys.exit(1)
    elif target.endswith(".txt"):
        try:
            lines = Path(target).read_text(encoding="utf-8").splitlines()
            for line in lines:
                line = line.strip()
                if line.startswith(("http://", "https://")):
                    posts.append(line)
            # Find common subdomain/username to name output directory
            if posts:
                username, _ = parse_url_target(posts[0])
                output_dir = Path("save_posts_output") / username
            else:
                output_dir = Path("save_posts_output") / "saved_posts"
            print(f"[bold $success]Loaded {len(posts)} posts from text file: {target} -> Saving under folder: {output_dir}[/bold $success]")
        except Exception as e:
            print(f"[bold red]Failed to read input file {target}: {e}[/bold red]")
            sys.exit(1)
    elif target.startswith(("http://", "https://")):
        posts = [target]
        username, _ = parse_url_target(target)
        output_dir = Path("save_posts_output") / username
        print(f"[bold $text-success]Saving single post URL: {target} -> Saving under folder: {output_dir}[/bold $text-success]\n")
    else:
        print(f"[bold $text-error]Invalid target '{target}'. Please specify a post URL, a .xlsx file, or a .txt file containing URLs.[/bold $text-error]")

    if not posts:
        print("[bold $text-error]No post URLs found to save.[/bold $text-error]")
        return

    update_status("Launching browser context...")
    async with async_playwright() as p:
        context = await launch_browser_with_fallback(
            p,
            user_data_dir=str(settings.get("user_data_dir", DEFAULT_USER_DATA_DIR)),
            headless=bool(settings.get("headless", True)),
            args=["--disable-dev-shm-usage"]
        )
        try:
            page = await context.new_page()
            # Set default timeouts
            timeout = settings.get("timeout", 30)
            page.set_default_timeout(int(timeout * 1000))
            page.set_default_navigation_timeout(int(timeout * 1000))
            results = await save_posts(page, posts, output_dir, delay=settings.get("delay", 0.0))
            print(f"\n[bold $success]Completed! Saved {results['success_count']} posts. Failed: {len(results['failed_urls'])}[/bold $success]")
        finally:
            await context.close()


def main_cli():
    try:
        asyncio.run(main_async())
    except KeyboardInterrupt:
        print("\n[bold red]Operation cancelled by user.[/bold red]")
        sys.exit(1)
    except Exception as e:
        if "AuthenticationError" in type(e).__name__:
            print(f"\n[bold red]❌ Error: Unable to download private photos/posts. {e}[/bold red]")
            print("[bold red]Please run 'lj-scraper --login' to authenticate first, or check your login session.[/bold red]")
            sys.exit(1)
        raise e


if __name__ == "__main__":
    main_cli()
